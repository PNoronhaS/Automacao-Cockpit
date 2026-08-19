import os
import time
import webbrowser
import pyautogui
import csv
import glob
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Font
import win32com.client as win32 

# ---------------- Configurações ----------------
URL_LOGIN = "https://sistema.exemplo.com.br/Account/Login"
DOWNLOADS_DIR = os.path.join(os.path.expanduser("~"), "Downloads")
DESTINO_DIR = r"C:\Users\USUARIO\Documents\Relatorios"

# ---------------- Coordenadas (1366x768) ----------------
COORD_USUARIO = (747, 402)
COORD_SENHA = (560, 484)
COORD_LOGIN = (687, 557)
COORD_ABA_DASHBOARD = (49, 342)
COORD_EXPORTAR = (353, 342)
COORD_EXPORTAR2 = (358, 300)   # coordenada alternativa
COORD_VOLTAR = (1302, 181)

# ---------------- Emails por loja ----------------
emails_por_loja = {
    "Loja A": "loja.a@exemplo.com",
    "Loja B": "loja.b@exemplo.com",
    "Loja C": "loja.c@exemplo.com"
}

# Alias para normalizar nomes vindos do CSV
alias_lojas = {
    "lojaa": "Loja A",
    "lojab": "Loja B"
}

# ---------------- Funções ----------------
def abrir_navegador_e_ir_para_login():
    print(">>> Abrindo navegador...")
    webbrowser.open(URL_LOGIN)
    time.sleep(10)

def fazer_login(usuario: str, senha: str):
    print(f">>> Fazendo login com usuário: {usuario}")
    pyautogui.click(*COORD_USUARIO)
    pyautogui.hotkey("ctrl", "a"); pyautogui.press("backspace"); pyautogui.typewrite(usuario)
    pyautogui.click(*COORD_SENHA)
    pyautogui.hotkey("ctrl", "a"); pyautogui.press("backspace"); pyautogui.typewrite(senha)
    pyautogui.click(*COORD_LOGIN)
    time.sleep(10)

def enviar_relatorio_por_email(loja_nome, arquivo_path):
    print(f">>> Enviando relatório da loja {loja_nome}...")
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)

    loja_nome_normalizado = alias_lojas.get(loja_nome.lower(), loja_nome)
    destinatario_loja = emails_por_loja.get(loja_nome_normalizado, None)

    if not destinatario_loja:
        print(f"!!! Nenhum e-mail configurado para a loja {loja_nome}.")
        return

    mail.To = destinatario_loja
    mail.Subject = f"Relatório Diário - {loja_nome_normalizado}"
    mail.Body = f"Segue em anexo o relatório diário da loja {loja_nome_normalizado}."
    mail.Attachments.Add(os.path.abspath(arquivo_path))
    mail.Send()
    print(f">>> E-mail enviado para {destinatario_loja}")

def exportar_relatorio(loja_nome: str):
    print(f">>> Exportando relatório da loja: {loja_nome}")
    pyautogui.click(*COORD_ABA_DASHBOARD)
    time.sleep(40)

    pyautogui.scroll(-500); time.sleep(1)
    pyautogui.scroll(-500); time.sleep(1)
    pyautogui.scroll(-500); time.sleep(1)

    pyautogui.click(*COORD_EXPORTAR)
    time.sleep(7)

    arquivos = glob.glob(os.path.join(DOWNLOADS_DIR, "*.xlsx"))

    if not arquivos:
        print(">>> Nenhum arquivo baixado, tentando coordenada alternativa...")
        pyautogui.click(*COORD_EXPORTAR2)
        time.sleep(7)
        arquivos = glob.glob(os.path.join(DOWNLOADS_DIR, "*.xlsx"))

    if arquivos:
        arquivo_recente = max(arquivos, key=os.path.getctime)
        data_atual = datetime.now().strftime("%d-%m")
        novo_nome = f"Relatorio_Diario_{loja_nome}_{data_atual}.xlsx"
        novo_caminho = os.path.join(DESTINO_DIR, novo_nome)
        os.makedirs(DESTINO_DIR, exist_ok=True)

        wb = openpyxl.load_workbook(arquivo_recente)
        ws = wb.active
        ws.delete_cols(2)
        ws.delete_cols(3)

        ws["A1"] = "Sistema - Restrito"
        ws.merge_cells("A1:B1")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].font = Font(size=14, bold=True)

        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A2"].font = Font(bold=True)
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B2"].font = Font(bold=True)

        for col in ["A", "B"]:
            for cell in ws[col]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 15
        ws.freeze_panes = "A3"

        wb.save(novo_caminho)
        os.remove(arquivo_recente)
        print(f">>> Arquivo salvo em: {novo_caminho}")

        enviar_relatorio_por_email(loja_nome, novo_caminho)
    else:
        print("!!! Nenhum arquivo .xlsx encontrado na pasta de downloads.")

    pyautogui.click(*COORD_VOLTAR)
    time.sleep(5)

def ler_lojas_csv(caminho='lojas.csv'):
    lojas = []
    try:
        with open(caminho, newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                lojas.append(row)
        print(f">>> {len(lojas)} lojas carregadas do CSV.")
    except FileNotFoundError:
        print("!!! Arquivo lojas.csv não encontrado.")
    return lojas

# ---------------- Execução principal ----------------
if __name__ == "__main__":
    pyautogui.FAILSAFE = True
    print(">>> Iniciando script...")
    lojas = ler_lojas_csv()

    if not lojas:
        print("!!! Nenhuma loja encontrada. Verifique o arquivo lojas.csv.")
    else:
        abrir_navegador_e_ir_para_login()
        for loja in lojas:
            print(f">>> Processando loja: {loja['loja']}")
            fazer_login(loja['usuario'], loja['senha'])
            exportar_relatorio(loja['loja'])
        print(">>> Todos os relatórios foram enviados com sucesso.")
